import openpyxl

book = openpyxl.Workbook()
print(book.sheetnames)
book.create_sheet('Estabelecimentos')
estab_page = book['Estabelecimentos']
estab_page.append(['ADQUIRENTES','DATA_INICIO','ESTABELECIMENTOS','CNPJ/CPF'])
estab_page.append(['PAGSEGURO','26/04/2022','321654','32.321.3231/0001-10'])
book.save('Planilha de Estabelecimentos.xlsx')