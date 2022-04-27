import openpyxl

book = openpyxl.Workbook()
print(book.sheetnames)
book.create_sheet('Frutas')
frutas_page = book['Frutas']
frutas_page.append(['Frutas','Quantidades','Preços'])
frutas_page.append(['Banana','5','R$3,90'])
book.save('Planilha de Compras.xlsx')